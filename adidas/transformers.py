from datetime import datetime

import pandas as pd
import pydash
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from adidas.utils import create_directory


def hyperlink_columns(name):
    data = {
        "table-of-contents": [1],
        "product-information": [3],
        "product-coordinates": [5, 6],
        "product-sizes": [],
        "product-technologies": [4],
        "product-reviews": [],
    }
    return data[name]


def column_width_definitions(name):
    data = {
        "table-of-contents": [12.5, 30],
        "product-information": [
            11,
            31,
            12,
            22,
            19.29,
            22.29,
            23.14,
            12.71,
            24.57,
            57.43,
            57.43,
            28.86,
            16.14,
            20.86,
            21.14,
            18,
            30.86,
            22.71,
            14.43,
        ],
        "product-coordinates": [15.43, 21.43, 30.29, 30.29, 27.86, 22.14, 56.43],
        "product-sizes": [11, 31, 13.5],
        "product-technologies": [11, 31, 19.14, 55, 55],
        "product-reviews": [11, 31, 13.57, 15.29, 29.29, 64, 19.57],
    }
    return data[name]


def format_sheet_views(writer, sheet_name, number_of_columns, name):
    column_widths = column_width_definitions(name)
    hyperlinks = hyperlink_columns(name)

    worksheet = writer.sheets[sheet_name]
    worksheet.freeze_panes = worksheet["A4"]

    for col in range(1, number_of_columns + 2):
        column_letter = get_column_letter(col)
        column_dimensions = worksheet.column_dimensions[column_letter]
        if sheet_name != "Sizes":
            column_dimensions.width = column_widths[col - 2] if col > 1 else 2.71
        else:
            if col == 1:
                column_dimensions.width = 2.71
            elif col == 2:
                column_dimensions.width = column_widths[0]
            elif col == 3:
                column_dimensions.width = column_widths[1]
            else:
                column_dimensions.width = column_widths[2]

        for cell in worksheet[column_letter]:
            if cell.row != 1:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.row > 3 and cell.column - 2 in hyperlinks:
                if name == "table-of-contents":
                    cell.value = f'=HYPERLINK("#\'{cell.value}\'!A1", "{cell.value}")'
                else:
                    cell.hyperlink = cell.value
                cell.font = Font(underline="single", color="0563C1")


def add_sheet_title(writer, sheet_name):
    worksheet = writer.sheets[sheet_name]
    cell = worksheet.cell(row=1, column=2)
    cell.value = f"{sheet_name} Table"
    cell.font = Font(color="002060", size=14, bold=True)


def format_column_headers(writer, sheet_name: str):
    worksheet = writer.sheets[sheet_name]
    worksheet.column_dimensions["A"].width = 2.71

    header_font = Font(color="002060", size=12, bold=True)
    header_alignment = Alignment(horizontal="left")
    header_border = Border(bottom=Side(border_style="medium", color="002060"))

    for col_idx, _ in enumerate(worksheet.columns):
        if col_idx == 0:
            continue

        cell = worksheet.cell(row=3, column=col_idx + 1)
        cell.value = pydash.human_case(cell.value).title()
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border


def create_contents_table_data(sheets):
    data = []
    for index, (key, sheet) in enumerate(sheets.items()):
        if key != "table-of-contents":
            data.append({"sheet_index": index, "sheet_name": sheet})
    return pd.DataFrame(data)


def generate_product_spreadsheet():
    destination = create_directory("data/spreadsheets", "xlsx")
    current_date = datetime.now().date().isoformat()
    writer = pd.ExcelWriter(f"{destination}/latest.xlsx", engine="openpyxl")

    sheets = {
        "table-of-contents": "Table of Contents",
        "product-information": "Product Information",
        "product-coordinates": "Coordinated Products",
        "product-sizes": "Sizes",
        "product-technologies": "Technologies",
        "product-reviews": "Reviews",
    }

    for filename, sheet_name in sheets.items():
        source = f"data/jsonlines/{current_date}/{filename}-latest.jl"
        if filename == "table-of-contents":
            df = create_contents_table_data(sheets)
        else:
            df = pd.read_json(source, lines=True)

        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2, startcol=1)

        # apply styles to excel sheet
        format_column_headers(writer, sheet_name)
        add_sheet_title(writer, sheet_name)
        format_sheet_views(writer, sheet_name, len(df.columns), filename)

    writer.close()
